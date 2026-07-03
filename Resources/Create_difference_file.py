import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
# from tkinter import filedialog as fd    
# from GeneratePDF import generate_pdf
import os
import sys

cwd = os.path.dirname(os.path.realpath(__file__))
sys.path.append(os.path.abspath(cwd))

def get_sort_value(d):
    val = list(d.values())[0][2]
    return abs(val) if isinstance(val, (int, float)) else -1

def compare_and_highlight(base_file, test_file, output_file):
    # Load both Excel files into pandas DataFrames
    df_main = pd.read_excel(base_file, sheet_name=None, index_col=None)

    df_test = pd.read_excel(test_file, sheet_name=None, index_col=None)

    # Load the test workbook to preserve formatting
    wb = load_workbook(test_file)
    
    # Define a red fill for highlighting
    red_fill = PatternFill(start_color='FF1F1F', end_color='FF1F1F', fill_type='solid')
    dark_orange_fill = PatternFill(start_color='DF5C16', end_color='DF5C16', fill_type='solid')
    orange_fill = PatternFill(start_color='FFB01F', end_color='FFB01F', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF33', end_color='FFFF33', fill_type='solid')
    light_yellow_fill = PatternFill(start_color='F9F9C8', end_color='F9F9C8', fill_type='solid')
    
    issue_exist = False

    all_differences = {}

    for sheet_name in df_main.keys():
        
        # Ensure the sheet exists in both files
        if sheet_name in df_test:
            differences_list = []
            df_main_sheet = df_main[sheet_name]
            df_test_sheet = df_test[sheet_name]
            ws = wb[sheet_name]

            # Ensure columns and indices match
            common_columns = df_main_sheet.columns.intersection(df_test_sheet.columns)
            common_indices = df_main_sheet.index.intersection(df_test_sheet.index)
            
            # Iterate over each cell to compare and mark differences
            for col in common_columns:
                for index in common_indices:
                    # print(index, col)
                    value_main = df_main_sheet.at[index, col]
                    value_test = df_test_sheet.at[index, col]

                    # Check if both values are non-empty and compare them
                    if pd.notna(value_main) and pd.notna(value_test) and value_main != value_test:

                        # Handle string comparison
                        if isinstance(value_main, (str,bool, np.bool_)) and isinstance(value_test, (str,bool, np.bool_)):
                            comment_text = (f"Main: {value_main}\nTest: {value_test}")
                            cell = ws.cell(row=index+2, column=df_main_sheet.columns.get_loc(col)+1)
                            cell.comment = Comment(comment_text, "Compare")
                            cell.fill = red_fill  # Mark all string differences in red
                            # differences_list.append({col : {value_main : value_test}})
                            issue_exist = True

                            differences_list.append({col: [value_test, value_main, "N/A"]})

                        # Handle numeric comparison
                        elif isinstance(value_main, (int, float, np.number)) and isinstance(value_test, (int, float, np.number)):
                            difference = round(value_test - value_main,6)
                            if value_main != 0:
                                percentage_difference = (difference / value_main) * 100
                                comment_text = (f"Difference: {difference:.4f}\n"
                                                f"Percentage: {percentage_difference:.4f}%\n"
                                                f"(Main: {value_main:.4f}, Test: {value_test:.4f})")
                            else:
                                percentage_difference = difference
                                comment_text = (f"Difference: {difference:.4f}\n"
                                                f"(Main: 0, Test: {value_test:.4f})")
                                
                            cell = ws.cell(row=index+2, column=df_main_sheet.columns.get_loc(col)+1)
                            cell.comment = Comment(comment_text, "Compare")
                            if percentage_difference > 50 or percentage_difference < -50:
                                cell.fill = red_fill
                                issue_exist = True
                            elif percentage_difference > 5 or percentage_difference < -5:
                                cell.fill = dark_orange_fill
                                issue_exist = True
                            elif percentage_difference > 1 or percentage_difference < -1:
                                cell.fill = orange_fill
                                issue_exist = True
                            elif percentage_difference > 0.1 or percentage_difference < -0.1:
                                cell.fill = yellow_fill
                                issue_exist = True
                            elif percentage_difference > 0.001 or percentage_difference < -0.001:
                                cell.fill = light_yellow_fill
                                issue_exist = True

                            if percentage_difference > 0.001 or percentage_difference < -0.001:    
                                differences_list.append({(col) : [float(value_main), float(value_test), round(float(percentage_difference), 3)]})

            if differences_list:
                # sorted_list = sorted(differences_list, key=lambda d: abs(list(d.values())[0][2]), reverse=True)
                sorted_list = sorted(differences_list, key=get_sort_value, reverse=True)
                all_differences[sheet_name] = sorted_list[:3] if len(sorted_list) >= 3 else sorted_list[:1]

            # if differences_list and not all(v[2] for v in differences_list.values()) == 0:
            #     top_three = sorted(v[2], reverse=True)[:3]
            #     if len(top_three) == 3:
            #         all_differences[sheet_name] = top_three
            #     else:
            #         all_differences[sheet_name] = max(differences_list)
            # if differences_list:
            #     sorted_list = sorted(
            #         differences_list,
            #         key=lambda x: abs(x["percentage"]),
            #         reverse=True
            #     )

            #     if len(sorted_list) >= 3:
            #         all_differences[sheet_name] = sorted_list[:3]
            #     else:
            #         all_differences[sheet_name] = [sorted_list[0]]

    if all_differences:            
        for key, val in all_differences.items():
            for item in val:
                for k, data in item.items():
                    if isinstance(data[2], (int, float)):
                        data[2] = f"{data[2]}%"
                    # data[2] = f"{data[2]}%"

    # Save the updated workbook with differences highlighted
    wb.save(output_file)
    # print(f"Differences saved to {output_file}")
    
    if issue_exist:
        return False, all_differences
    else:
        return True, all_differences

# base_file_path = fd.askopenfilename(title='Select the Base version file', filetypes=(('Excel files', '*.xlsx'),))
# test_file_path = fd.askopenfilename(title='Select the Test version file', filetypes=(('Excel files', '*.xlsx'),))
# test_file_name = os.path.basename(test_file_path)
# base_file_name = os.path.basename(base_file_path)
# print(test_file_name)
# print(base_file_name)
# output_folder = os.path.dirname(r"C:\Users\ruben.pulayath\OneDrive - TNEI Services Ltd\Desktop\Testing IPSA Akshay\Ruben Comparison Script\IEEE 9 Bus Network_LoadFlow_v3.2.0.xlsx",)

# output_file = os.path.join(output_folder, 'Test_differences.xlsx')
# comparison_success = compare_and_highlight(base_file_path, test_file_path, output_file)
# report_file_name = 'Comparison_Report.pdf'
# generate_pdf(report_file_name, base_file_name, test_file_name, comparison_success)
# compare_and_highlight(r"C:\Users\ruben.pulayath\OneDrive - TNEI Services Ltd\Desktop\Testing IPSA Akshay\Ruben Comparison Script\IEEE 9 Bus Network_LoadFlow_v3.1.0.xlsx", 
#                       r"C:\Users\ruben.pulayath\OneDrive - TNEI Services Ltd\Desktop\Testing IPSA Akshay\Ruben Comparison Script\IEEE 9 Bus Network_LoadFlow_v3.2.0.xlsx", 
#                       output_file)